from flask import request, jsonify, send_file
import io
import zipfile
from openpyxl.reader.excel import load_workbook
from ..openai_processing import extract_content, identify_requirements, apply_wrap_text, identify_page_limits, break_text_into_lines


def upload_files():
    # Expecting a .zip file
    zip_file = request.files.get('zip_file')

    if not zip_file:
        return jsonify({'error': 'Missing zip file!'}), 400

    try:
        # Create an in-memory buffer to hold the zip file contents
        zip_buffer = io.BytesIO(zip_file.read())

        # Open the zip file and extract its contents
        with zipfile.ZipFile(zip_buffer, 'r') as z:
            word_file = None
            excel_file = None

            # Loop through the files inside the zip and find the Word and Excel files
            for file_name in z.namelist():
                if file_name.endswith('.docx') or file_name.endswith('.doc'):
                    word_file = io.BytesIO(z.read(file_name))
                elif file_name.endswith('.xlsx') or file_name.endswith('.xls'):
                    excel_file = io.BytesIO(z.read(file_name))

            # Check if both files were found
            if not word_file or not excel_file:
                return jsonify({'error': 'Both Word and Excel files are required inside the zip.'}), 400

            # Load the Excel file
            wb = load_workbook(excel_file)
            ws = wb.active

            # Assume headers are already in the Excel template
            row_num = 2  # Start writing data from the second row

            # Extract content from the Word document using your extraction function
            sections = extract_content(word_file)

            for header, data in sections.items():
                content = data['content']
                page_limit = data.get('page_limit', 'Not specified')  # Default if no page limit specified

                # Use OpenAI to identify requirements for the given header and content
                requirements = identify_requirements(header, content)
                detected_page_limit = identify_page_limits(content)  # Assume content is string; adjust if otherwise

                # Write data to Excel using text breaking and applying wrap text
                ws[f"A{row_num}"].value = break_text_into_lines(header, 50)
                ws[f"B{row_num}"].value = break_text_into_lines(requirements, 50)
                ws[f"C{row_num}"].value = page_limit if page_limit else detected_page_limit

                apply_wrap_text(ws[f"A{row_num}"])
                apply_wrap_text(ws[f"B{row_num}"])
                apply_wrap_text(ws[f"C{row_num}"])

                row_num += 1

            # Create an in-memory buffer to store the modified Excel file
            excel_output = io.BytesIO()
            wb.save(excel_output)
            excel_output.seek(0)  # Reset the pointer to the start of the buffer

            # Create an in-memory zip archive and add the processed Excel file to it
            zip_output = io.BytesIO()
            with zipfile.ZipFile(zip_output, 'w', zipfile.ZIP_DEFLATED) as zipf:
                zipf.writestr('processed_result.xlsx', excel_output.getvalue())

            zip_output.seek(0)  # Reset the pointer to the start of the buffer

            # Send the zip file back to the client
            return send_file(
                zip_output,
                as_attachment=True,
                download_name='processed_files.zip',
                mimetype='application/zip'
            )

    except Exception as e:
        print(e)
        return jsonify({'error': f'An error occurred during processing: {str(e)}'}), 500
